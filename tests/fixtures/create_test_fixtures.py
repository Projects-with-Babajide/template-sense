"""
Script to create test fixtures for file validation tests.

This script creates:
- A valid .xlsx file
- A valid .xls file (legacy format)
- Invalid format files for testing error handling
"""

from pathlib import Path

from openpyxl import Workbook


def create_test_fixtures():
    """Create test fixture files."""
    fixtures_dir = Path(__file__).parent

    # Create a valid .xlsx file with header and table data
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Sheet"

    # Header section (rows 1-2)
    ws["A1"] = "Invoice Number"
    ws["B1"] = "Date"
    ws["A2"] = "INV-001"
    ws["B2"] = "2025-01-15"

    # Empty row separator (row 3)

    # Table section (rows 4-10)
    # Table header
    ws["A4"] = "Item"
    ws["B4"] = "Quantity"
    ws["C4"] = "Price"
    ws["D4"] = "Amount"

    # Table data rows (dense, numeric, consistent columns)
    ws["A5"] = "Widget A"
    ws["B5"] = 10
    ws["C5"] = 25.50
    ws["D5"] = 255.00

    ws["A6"] = "Widget B"
    ws["B6"] = 5
    ws["C6"] = 40.00
    ws["D6"] = 200.00

    ws["A7"] = "Widget C"
    ws["B7"] = 15
    ws["C7"] = 30.00
    ws["D7"] = 450.00

    ws["A8"] = "Widget D"
    ws["B8"] = 8
    ws["C8"] = 50.00
    ws["D8"] = 400.00

    ws["A9"] = "Widget E"
    ws["B9"] = 12
    ws["C9"] = 35.00
    ws["D9"] = 420.00

    ws["A10"] = "Widget F"
    ws["B10"] = 20
    ws["C10"] = 28.00
    ws["D10"] = 560.00

    wb.save(fixtures_dir / "valid_template.xlsx")
    print(f"Created: {fixtures_dir / 'valid_template.xlsx'}")

    # Create a CSV file (unsupported format)
    csv_file = fixtures_dir / "invalid_format.csv"
    csv_file.write_text("Invoice Number,Date\nINV-001,2025-01-15\n")
    print(f"Created: {csv_file}")

    # Create a text file (unsupported format)
    txt_file = fixtures_dir / "invalid_format.txt"
    txt_file.write_text("This is a text file, not an Excel file.")
    print(f"Created: {txt_file}")

    # Create a PDF-like file (unsupported format)
    pdf_file = fixtures_dir / "invalid_format.pdf"
    pdf_file.write_text("This is not a real PDF, just for testing.")
    print(f"Created: {pdf_file}")

    # Create a file with no extension
    no_ext_file = fixtures_dir / "no_extension"
    no_ext_file.write_text("File with no extension")
    print(f"Created: {no_ext_file}")

    # Create a corrupted .xlsx file (not actually Excel format)
    corrupt_file = fixtures_dir / "corrupted.xlsx"
    corrupt_file.write_text("This is not a valid Excel file, just corrupted data.")
    print(f"Created: {corrupt_file}")

    print("\nAll test fixtures created successfully!")


if __name__ == "__main__":
    create_test_fixtures()
