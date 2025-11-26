"""
Unit tests for sheet_extractor module.

Tests the sheet-level extraction utilities including:
- Extracting raw grids from sheets
- Filtering non-empty rows
- Filtering non-empty columns
- Calculating used ranges
- Error handling for invalid sheets
- Ensuring no openpyxl objects are exposed
"""

from pathlib import Path

import pytest
from openpyxl import Workbook

from template_sense.adapters.excel_adapter import ExcelWorkbook
from template_sense.errors import ExtractionError
from template_sense.extraction.sheet_extractor import (
    extract_non_empty_columns,
    extract_non_empty_rows,
    extract_raw_grid,
    get_used_range,
)
from template_sense.file_loader import load_excel_file

# Fixtures


@pytest.fixture
def simple_grid_workbook(tmp_path: Path) -> tuple[Path, Workbook]:
    """Create a workbook with a simple 3x3 grid."""
    file_path = tmp_path / "simple_grid.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SimpleGrid"

    # Create 3x3 grid
    sheet["A1"] = "Header1"
    sheet["B1"] = "Header2"
    sheet["C1"] = "Header3"
    sheet["A2"] = "Value1"
    sheet["B2"] = "Value2"
    sheet["C2"] = "Value3"
    sheet["A3"] = "Value4"
    sheet["B3"] = "Value5"
    sheet["C3"] = "Value6"

    workbook.save(file_path)
    raw_workbook = load_excel_file(file_path)

    yield file_path, raw_workbook

    raw_workbook.close()


@pytest.fixture
def empty_sheet_workbook(tmp_path: Path) -> tuple[Path, Workbook]:
    """Create a workbook with an empty sheet."""
    file_path = tmp_path / "empty_sheet.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "EmptySheet"
    # Don't add any data

    workbook.save(file_path)
    raw_workbook = load_excel_file(file_path)

    yield file_path, raw_workbook

    raw_workbook.close()


@pytest.fixture
def trailing_empty_workbook(tmp_path: Path) -> tuple[Path, Workbook]:
    """Create a workbook with data surrounded by empty rows/columns."""
    file_path = tmp_path / "trailing_empty.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "TrailingEmpty"

    # Add data in middle with empty borders
    sheet["B2"] = "A"
    sheet["C2"] = "B"
    sheet["B3"] = "C"
    sheet["C3"] = "D"

    # Leave row 1, column A, and rows/columns after C3 empty

    workbook.save(file_path)
    raw_workbook = load_excel_file(file_path)

    yield file_path, raw_workbook

    raw_workbook.close()


@pytest.fixture
def mixed_empty_workbook(tmp_path: Path) -> tuple[Path, Workbook]:
    """Create a workbook with mixed empty and non-empty rows/columns."""
    file_path = tmp_path / "mixed_empty.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MixedEmpty"

    # Row 1: has data
    sheet["A1"] = "Header1"
    sheet["B1"] = "Header2"

    # Row 2: empty

    # Row 3: has data
    sheet["A3"] = "Value1"
    sheet["B3"] = "Value2"

    # Column C: empty in all rows

    # Column D: has data
    sheet["D1"] = "Col4"
    sheet["D3"] = "Val4"

    workbook.save(file_path)
    raw_workbook = load_excel_file(file_path)

    yield file_path, raw_workbook

    raw_workbook.close()


@pytest.fixture
def merged_cells_workbook(tmp_path: Path) -> tuple[Path, Workbook]:
    """Create a workbook with merged cells."""
    file_path = tmp_path / "merged_cells.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MergedCells"

    # Add merged cells
    sheet.merge_cells("A1:B1")
    sheet["A1"] = "Merged Header"

    sheet["A2"] = "Data1"
    sheet["B2"] = "Data2"

    workbook.save(file_path)
    raw_workbook = load_excel_file(file_path)

    yield file_path, raw_workbook

    raw_workbook.close()


# Tests for extract_raw_grid


def test_extract_raw_grid_simple(simple_grid_workbook):
    """Test extracting a simple 3x3 grid."""
    _, raw_workbook = simple_grid_workbook
    wb = ExcelWorkbook(raw_workbook)

    grid = extract_raw_grid(wb, "SimpleGrid")

    assert isinstance(grid, list)
    assert len(grid) == 3
    assert grid[0] == ["Header1", "Header2", "Header3"]
    assert grid[1] == ["Value1", "Value2", "Value3"]
    assert grid[2] == ["Value4", "Value5", "Value6"]


def test_extract_raw_grid_empty_sheet(empty_sheet_workbook):
    """Test extracting from an empty sheet."""
    _, raw_workbook = empty_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    grid = extract_raw_grid(wb, "EmptySheet")

    assert isinstance(grid, list)
    # Empty sheets may have one empty row from openpyxl
    # We just verify it's a list and doesn't crash


def test_extract_raw_grid_returns_primitives(simple_grid_workbook):
    """Test that extract_raw_grid returns primitive values, not Cell objects."""
    _, raw_workbook = simple_grid_workbook
    wb = ExcelWorkbook(raw_workbook)

    grid = extract_raw_grid(wb, "SimpleGrid")

    for row in grid:
        assert isinstance(row, list)
        for cell in row:
            # Should be primitive types (str, int, float, None, etc.)
            # Not openpyxl Cell objects (which have .value attribute)
            assert not hasattr(cell, "value")


def test_extract_raw_grid_invalid_sheet_raises_error(simple_grid_workbook):
    """Test that extracting from non-existent sheet raises ExtractionError."""
    _, raw_workbook = simple_grid_workbook
    wb = ExcelWorkbook(raw_workbook)

    with pytest.raises(ExtractionError) as exc_info:
        extract_raw_grid(wb, "NonExistentSheet")

    assert "does not exist" in str(exc_info.value).lower()


def test_extract_raw_grid_with_trailing_empty(trailing_empty_workbook):
    """Test extracting grid with trailing empty rows/columns."""
    _, raw_workbook = trailing_empty_workbook
    wb = ExcelWorkbook(raw_workbook)

    grid = extract_raw_grid(wb, "TrailingEmpty")

    # Grid includes empty rows/columns as openpyxl sees them
    assert isinstance(grid, list)
    assert len(grid) >= 3  # Should have at least 3 rows


def test_extract_raw_grid_with_merged_cells(merged_cells_workbook):
    """Test extracting grid with merged cells."""
    _, raw_workbook = merged_cells_workbook
    wb = ExcelWorkbook(raw_workbook)

    grid = extract_raw_grid(wb, "MergedCells")

    assert isinstance(grid, list)
    assert len(grid) >= 2
    # Merged cells: only top-left cell has value, others are None
    assert grid[0][0] == "Merged Header"


# Tests for extract_non_empty_rows


def test_extract_non_empty_rows_basic():
    """Test filtering non-empty rows from a grid."""
    grid = [
        ["Header1", "Header2"],
        ["Value1", "Value2"],
        [None, None],
        ["Value3", "Value4"],
        ["", ""],
    ]

    result = extract_non_empty_rows(grid)

    assert len(result) == 3
    assert result[0] == ["Header1", "Header2"]
    assert result[1] == ["Value1", "Value2"]
    assert result[2] == ["Value3", "Value4"]


def test_extract_non_empty_rows_all_empty():
    """Test filtering when all rows are empty."""
    grid = [[None, None], ["", ""], [None, ""]]

    result = extract_non_empty_rows(grid)

    assert result == []


def test_extract_non_empty_rows_all_non_empty():
    """Test filtering when all rows have data."""
    grid = [["A", "B"], ["C", "D"], ["E", "F"]]

    result = extract_non_empty_rows(grid)

    assert len(result) == 3
    assert result == grid


def test_extract_non_empty_rows_partial_empty_cells():
    """Test that rows with at least one non-empty cell are kept."""
    grid = [
        ["A", None],
        [None, "B"],
        [None, None],
        ["", "C"],
    ]

    result = extract_non_empty_rows(grid)

    assert len(result) == 3
    assert result[0] == ["A", None]
    assert result[1] == [None, "B"]
    assert result[2] == ["", "C"]


def test_extract_non_empty_rows_empty_grid():
    """Test filtering an empty grid."""
    grid = []

    result = extract_non_empty_rows(grid)

    assert result == []


# Tests for extract_non_empty_columns


def test_extract_non_empty_columns_basic():
    """Test filtering non-empty columns from a grid."""
    grid = [
        ["A", None, "B", ""],
        ["C", None, "D", ""],
    ]

    result = extract_non_empty_columns(grid)

    assert len(result) == 2
    assert result[0] == ["A", "B"]
    assert result[1] == ["C", "D"]


def test_extract_non_empty_columns_all_empty():
    """Test filtering when all columns are empty."""
    grid = [
        [None, "", None],
        [None, "", None],
    ]

    result = extract_non_empty_columns(grid)

    assert result == [[], []]


def test_extract_non_empty_columns_all_non_empty():
    """Test filtering when all columns have data."""
    grid = [["A", "B", "C"], ["D", "E", "F"]]

    result = extract_non_empty_columns(grid)

    assert len(result) == 2
    assert result[0] == ["A", "B", "C"]
    assert result[1] == ["D", "E", "F"]


def test_extract_non_empty_columns_partial_empty_cells():
    """Test that columns with at least one non-empty cell are kept."""
    grid = [
        ["A", None, "", "D"],
        [None, None, "C", "E"],
    ]

    result = extract_non_empty_columns(grid)

    # Column 0: ["A", None] - has A, keep
    # Column 1: [None, None] - all empty, remove
    # Column 2: ["", "C"] - has C, keep
    # Column 3: ["D", "E"] - has both, keep
    # Result should have 3 columns: 0, 2, 3
    assert len(result) == 2  # 2 rows
    assert len(result[0]) == 3  # 3 columns
    assert result[0] == ["A", "", "D"]
    assert result[1] == [None, "C", "E"]


def test_extract_non_empty_columns_empty_grid():
    """Test filtering an empty grid."""
    grid = []

    result = extract_non_empty_columns(grid)

    assert result == []


def test_extract_non_empty_columns_inconsistent_row_lengths():
    """Test filtering with inconsistent row lengths."""
    grid = [
        ["A", "B", "C"],
        ["D"],
        ["E", "F"],
    ]

    result = extract_non_empty_columns(grid)

    # All columns have at least one non-empty value
    assert len(result) == 3
    assert result[0] == ["A", "B", "C"]
    assert result[1] == ["D", None, None]  # Padded with None
    assert result[2] == ["E", "F", None]  # Padded with None


def test_extract_non_empty_columns_empty_rows():
    """Test filtering when grid has empty row lists."""
    grid = [[], [], []]

    result = extract_non_empty_columns(grid)

    assert result == []


# Tests for get_used_range


def test_get_used_range_simple(simple_grid_workbook):
    """Test getting used range for a simple 3x3 grid."""
    _, raw_workbook = simple_grid_workbook
    wb = ExcelWorkbook(raw_workbook)

    min_row, max_row, min_col, max_col = get_used_range(wb, "SimpleGrid")

    assert min_row == 1
    assert max_row == 3
    assert min_col == 1
    assert max_col == 3


def test_get_used_range_empty_sheet(empty_sheet_workbook):
    """Test getting used range for an empty sheet."""
    _, raw_workbook = empty_sheet_workbook
    wb = ExcelWorkbook(raw_workbook)

    min_row, max_row, min_col, max_col = get_used_range(wb, "EmptySheet")

    # Empty sheets return (1, 1, 1, 1)
    assert min_row == 1
    assert max_row == 1
    assert min_col == 1
    assert max_col == 1


def test_get_used_range_with_trailing_empty(trailing_empty_workbook):
    """Test getting used range excluding empty borders."""
    _, raw_workbook = trailing_empty_workbook
    wb = ExcelWorkbook(raw_workbook)

    min_row, max_row, min_col, max_col = get_used_range(wb, "TrailingEmpty")

    # Data is in B2:C3, so:
    # min_row=2, max_row=3, min_col=2, max_col=3
    assert min_row == 2
    assert max_row == 3
    assert min_col == 2
    assert max_col == 3


def test_get_used_range_with_mixed_empty(mixed_empty_workbook):
    """Test getting used range with mixed empty rows/columns."""
    _, raw_workbook = mixed_empty_workbook
    wb = ExcelWorkbook(raw_workbook)

    min_row, max_row, min_col, max_col = get_used_range(wb, "MixedEmpty")

    # Data spans rows 1-3 and columns 1,2,4 (A,B,D)
    # min_row=1, max_row=3, min_col=1, max_col=4
    assert min_row == 1
    assert max_row == 3
    assert min_col == 1
    assert max_col == 4


def test_get_used_range_returns_1_based_indices(simple_grid_workbook):
    """Test that get_used_range returns 1-based indices (Excel convention)."""
    _, raw_workbook = simple_grid_workbook
    wb = ExcelWorkbook(raw_workbook)

    min_row, max_row, min_col, max_col = get_used_range(wb, "SimpleGrid")

    # All indices should be >= 1 (1-based)
    assert min_row >= 1
    assert max_row >= 1
    assert min_col >= 1
    assert max_col >= 1


def test_get_used_range_invalid_sheet_raises_error(simple_grid_workbook):
    """Test that getting used range for non-existent sheet raises ExtractionError."""
    _, raw_workbook = simple_grid_workbook
    wb = ExcelWorkbook(raw_workbook)

    with pytest.raises(ExtractionError) as exc_info:
        get_used_range(wb, "NonExistentSheet")

    assert "does not exist" in str(exc_info.value).lower()


def test_get_used_range_single_cell(tmp_path: Path):
    """Test getting used range for a sheet with single cell."""
    file_path = tmp_path / "single_cell.xlsx"

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "SingleCell"
    sheet["C3"] = "SingleValue"

    workbook.save(file_path)
    raw_workbook = load_excel_file(file_path)
    wb = ExcelWorkbook(raw_workbook)

    min_row, max_row, min_col, max_col = get_used_range(wb, "SingleCell")

    # Single cell at C3 (row=3, col=3)
    assert min_row == 3
    assert max_row == 3
    assert min_col == 3
    assert max_col == 3

    raw_workbook.close()


# Integration tests


def test_full_extraction_workflow(mixed_empty_workbook):
    """Test a complete extraction workflow with all utilities."""
    _, raw_workbook = mixed_empty_workbook
    wb = ExcelWorkbook(raw_workbook)

    # 1. Extract raw grid
    grid = extract_raw_grid(wb, "MixedEmpty")
    assert len(grid) >= 3

    # 2. Filter non-empty rows
    non_empty_rows = extract_non_empty_rows(grid)
    assert len(non_empty_rows) == 2  # Rows 1 and 3 have data

    # 3. Filter non-empty columns
    non_empty_cols = extract_non_empty_columns(grid)
    # Columns A, B, D have data (column C is empty)
    assert all(len(row) == 3 for row in non_empty_cols)

    # 4. Get used range
    min_row, max_row, min_col, max_col = get_used_range(wb, "MixedEmpty")
    assert min_row == 1
    assert max_row == 3
    assert min_col == 1
    assert max_col == 4


def test_no_openpyxl_objects_in_results(simple_grid_workbook):
    """Test that no openpyxl objects leak into any extraction results."""
    _, raw_workbook = simple_grid_workbook
    wb = ExcelWorkbook(raw_workbook)

    # Extract raw grid
    grid = extract_raw_grid(wb, "SimpleGrid")
    for row in grid:
        for cell in row:
            assert not hasattr(cell, "value")  # No Cell objects

    # Filter non-empty rows
    non_empty_rows = extract_non_empty_rows(grid)
    for row in non_empty_rows:
        for cell in row:
            assert not hasattr(cell, "value")

    # Filter non-empty columns
    non_empty_cols = extract_non_empty_columns(grid)
    for row in non_empty_cols:
        for cell in row:
            assert not hasattr(cell, "value") if cell is not None else True

    # Get used range returns tuple of ints
    result = get_used_range(wb, "SimpleGrid")
    assert isinstance(result, tuple)
    assert all(isinstance(val, int) for val in result)


def test_extraction_with_valid_template_fixture():
    """Test extraction utilities with the existing valid_template.xlsx fixture."""
    file_path = Path("tests/fixtures/valid_template.xlsx")
    raw_workbook = load_excel_file(file_path)
    wb = ExcelWorkbook(raw_workbook)

    # Extract raw grid
    grid = extract_raw_grid(wb, "Test Sheet")
    assert len(grid) >= 2
    assert grid[0][0] == "Invoice Number"

    # Filter non-empty rows
    non_empty_rows = extract_non_empty_rows(grid)
    assert len(non_empty_rows) >= 2

    # Get used range
    min_row, max_row, min_col, max_col = get_used_range(wb, "Test Sheet")
    assert min_row >= 1
    assert max_row >= 2
    assert min_col >= 1
    assert max_col >= 2

    raw_workbook.close()
