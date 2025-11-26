"""
Excel workbook abstraction layer for Template Sense.

This module provides a clean interface for working with Excel workbooks that shields
the extraction pipeline from openpyxl internals. This abstraction enables future support
for additional file formats (PDF, DOCX, CSV) using the same adapter pattern.

Classes:
    ExcelWorkbook: Abstraction layer wrapping openpyxl Workbook objects

Usage Example:
    from pathlib import Path
    from template_sense.file_loader import load_excel_file
    from template_sense.adapters.excel_adapter import ExcelWorkbook

    # Load workbook using file_loader
    raw_workbook = load_excel_file(Path("template.xlsx"))

    # Wrap in abstraction layer
    workbook = ExcelWorkbook(raw_workbook)

    # Use clean interface
    sheet_names = workbook.get_sheet_names()
    for row_values in workbook.iter_rows("Sheet1"):
        print(row_values)

    workbook.close()
"""

import logging
from collections.abc import Generator
from typing import Any

from openpyxl.workbook.workbook import Workbook

from template_sense.errors import ExtractionError, FileValidationError

# Set up module logger
logger = logging.getLogger(__name__)


class ExcelWorkbook:
    """
    Abstraction layer for Excel workbooks.

    Wraps openpyxl Workbook and provides a clean interface for the extraction pipeline.
    This class never exposes openpyxl objects directly to consumers, ensuring that
    future extraction logic can remain independent of the underlying Excel library.

    Args:
        workbook: An openpyxl Workbook object (typically from file_loader.load_excel_file)

    Example:
        >>> from template_sense.file_loader import load_excel_file
        >>> from pathlib import Path
        >>> raw_wb = load_excel_file(Path("template.xlsx"))
        >>> wb = ExcelWorkbook(raw_wb)
        >>> print(wb.get_sheet_names())
        ['Sheet1', 'Invoice']
        >>> wb.close()
    """

    def __init__(self, workbook: Workbook):
        """
        Initialize the ExcelWorkbook abstraction.

        Args:
            workbook: An openpyxl Workbook object

        Raises:
            FileValidationError: If workbook is None or invalid
        """
        if workbook is None:
            raise FileValidationError(reason="Workbook cannot be None")

        # Store as private attribute to prevent direct access
        self._workbook = workbook
        logger.debug("ExcelWorkbook initialized with %d sheets", len(workbook.sheetnames))

    def get_sheet_names(self) -> list[str]:
        """
        Get list of all sheet names in the workbook.

        Returns:
            List of sheet names in the order they appear in the workbook

        Example:
            >>> wb.get_sheet_names()
            ['Sheet1', 'Invoice', 'Summary']
        """
        sheet_names = self._workbook.sheetnames
        logger.debug("Retrieved %d sheet names", len(sheet_names))
        return sheet_names

    def iter_rows(
        self,
        sheet_name: str,
        min_row: int = 1,
        max_row: int | None = None,
    ) -> Generator[list[Any], None, None]:
        """
        Iterate over rows in a sheet, returning cell values (not openpyxl Cell objects).

        Args:
            sheet_name: Name of the sheet to iterate over
            min_row: Starting row (1-based, Excel convention). Defaults to 1.
            max_row: Ending row (1-based, inclusive). If None, iterates to end of sheet.

        Yields:
            List of cell values for each row (primitives like str, int, float, None, etc.)

        Raises:
            ExtractionError: If sheet_name does not exist

        Example:
            >>> for row_values in wb.iter_rows("Sheet1"):
            ...     print(row_values)
            ['Header1', 'Header2']
            ['Value1', 'Value2']
        """
        # Validate sheet exists
        if sheet_name not in self._workbook.sheetnames:
            logger.error("Sheet '%s' not found in workbook", sheet_name)
            raise ExtractionError(
                extraction_type="sheet",
                reason=f"Sheet '{sheet_name}' does not exist",
            )

        sheet = self._workbook[sheet_name]
        logger.debug(
            "Iterating rows in sheet '%s' from row %d to %s",
            sheet_name,
            min_row,
            max_row if max_row else "end",
        )

        # Use openpyxl's iter_rows with values_only=True to get primitives, not Cell objects
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True):
            # Convert tuple to list for consistency
            yield list(row)

    def get_cell_value(self, sheet_name: str, row: int, col: int) -> Any:
        """
        Get value of a specific cell.

        Args:
            sheet_name: Name of the sheet
            row: 1-based row index (Excel convention)
            col: 1-based column index (Excel convention)

        Returns:
            Cell value (str, int, float, None, datetime, etc.)

        Raises:
            ExtractionError: If sheet_name does not exist or cell access fails

        Example:
            >>> wb.get_cell_value("Sheet1", 1, 1)
            'Header1'
            >>> wb.get_cell_value("Sheet1", 2, 2)
            'Value2'
        """
        # Validate sheet exists
        if sheet_name not in self._workbook.sheetnames:
            logger.error("Sheet '%s' not found in workbook", sheet_name)
            raise ExtractionError(
                extraction_type="sheet",
                reason=f"Sheet '{sheet_name}' does not exist",
            )

        try:
            sheet = self._workbook[sheet_name]
            # Access cell using 1-based indexing (openpyxl uses 1-based)
            cell_value = sheet.cell(row=row, column=col).value
            logger.debug(
                "Retrieved cell value at %s[%d, %d]: %s",
                sheet_name,
                row,
                col,
                repr(cell_value),
            )
            return cell_value

        except Exception as e:
            logger.error(
                "Failed to access cell at %s[%d, %d]: %s",
                sheet_name,
                row,
                col,
                str(e),
            )
            raise ExtractionError(
                extraction_type="cell",
                reason=f"Failed to access cell at row {row}, col {col}",
                row_index=row,
            ) from e

    def get_row_count(self, sheet_name: str) -> int:
        """
        Get total number of rows in a sheet.

        Args:
            sheet_name: Name of the sheet

        Returns:
            Number of rows (0 if sheet is empty)

        Raises:
            ExtractionError: If sheet_name does not exist

        Example:
            >>> wb.get_row_count("Sheet1")
            10
        """
        # Validate sheet exists
        if sheet_name not in self._workbook.sheetnames:
            logger.error("Sheet '%s' not found in workbook", sheet_name)
            raise ExtractionError(
                extraction_type="sheet",
                reason=f"Sheet '{sheet_name}' does not exist",
            )

        sheet = self._workbook[sheet_name]
        # max_row can be None for empty sheets
        row_count = sheet.max_row if sheet.max_row is not None else 0
        logger.debug("Sheet '%s' has %d rows", sheet_name, row_count)
        return row_count

    def get_column_count(self, sheet_name: str) -> int:
        """
        Get total number of columns in a sheet.

        Args:
            sheet_name: Name of the sheet

        Returns:
            Number of columns (0 if sheet is empty)

        Raises:
            ExtractionError: If sheet_name does not exist

        Example:
            >>> wb.get_column_count("Sheet1")
            5
        """
        # Validate sheet exists
        if sheet_name not in self._workbook.sheetnames:
            logger.error("Sheet '%s' not found in workbook", sheet_name)
            raise ExtractionError(
                extraction_type="sheet",
                reason=f"Sheet '{sheet_name}' does not exist",
            )

        sheet = self._workbook[sheet_name]
        # max_column can be None for empty sheets
        col_count = sheet.max_column if sheet.max_column is not None else 0
        logger.debug("Sheet '%s' has %d columns", sheet_name, col_count)
        return col_count

    def close(self) -> None:
        """
        Close the underlying workbook to free resources.

        It's good practice to call this when done with the workbook,
        though Python's garbage collector will also clean up eventually.

        Example:
            >>> wb = ExcelWorkbook(raw_workbook)
            >>> # ... use workbook ...
            >>> wb.close()
        """
        if self._workbook:
            self._workbook.close()
            logger.debug("Workbook closed")


__all__ = ["ExcelWorkbook"]
